
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import SerEvaluationForm
from .models import SerEvaluation, Negotiator

@login_required
def start_ser_evaluation_view(request, cedula):
    if request.user.role != 'lider':
        return redirect('profile')
    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    if request.method == 'POST':
        form = SerEvaluationForm(request.POST)
        if form.is_valid():
            SerEvaluation.objects.create(
                negotiator=negotiator,
                evaluator=request.user,
                actitud=form.cleaned_data['actitud'],
                trabajo_en_equipo=form.cleaned_data['trabajo_en_equipo'],
                sentido_pertenencia=form.cleaned_data['sentido_pertenencia'],
                relacionamiento=form.cleaned_data['relacionamiento'],
                compromiso=form.cleaned_data['compromiso'],
            )
            messages.success(request, '¡Evaluación del Ser registrada correctamente!')
            return redirect('negotiator_detail', cedula=cedula)
    else:
        form = SerEvaluationForm()
    return render(request, 'accounts/start_ser_evaluation.html', {
        'negotiator': negotiator,
        'form': form
    })
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from datetime import datetime, timedelta
import json
from django.conf import settings
from .models import Negotiator, Evaluation, KPI, EvaluationKPI, NegotiatorIndicator
from .forms import EvaluationForm
from django.db.models import Avg, Count
from django.db.models.functions import Coalesce
from django.contrib.auth import get_user_model

def home_view(request):
    return render(request, 'accounts/home.html')

@login_required
def profile_view(request):
    if request.user.role == 'lider':
        return redirect('lider_dashboard')
    else:
        return redirect('administrativo_dashboard')

@login_required
def lider_dashboard_view(request):
    negotiators = Negotiator.objects.filter(leader=request.user)
    pending_negotiators = request.user.get_negotiators_with_pending_evaluations()
    
    context = {
        'negotiators': negotiators,
        'pending_negotiators': pending_negotiators,
        'total_pending': len(pending_negotiators)
    }
    return render(request, 'accounts/lider_dashboard.html', context)

@login_required
def pending_evaluations_view(request):
    """
    Vista para mostrar las evaluaciones pendientes del líder
    """
    if request.user.role != 'lider':
        return redirect('profile')
    
    pending_negotiators = request.user.get_negotiators_with_pending_evaluations()
    
    context = {
        'pending_negotiators': pending_negotiators,
        'total_pending': len(pending_negotiators)
    }
    return render(request, 'accounts/pending_evaluations.html', context)

@login_required
def last_evaluation_view(request, cedula):
    """
    Vista para mostrar la última evaluación de un negociador
    """
    if request.user.role != 'lider':
        return redirect('profile')
    
    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    last_evaluation = negotiator.get_last_evaluation()
    
    evaluation_kpis = last_evaluation.kpis.all() if last_evaluation else []
    context = {
        'negotiator': negotiator,
        'last_evaluation': last_evaluation,
        'has_evaluations': negotiator.has_evaluations(),
        'evaluation_count': negotiator.get_evaluation_count(),
        'evaluation_kpis': evaluation_kpis
    }
    return render(request, 'accounts/last_evaluation.html', context)

@login_required
def start_evaluation_view(request, cedula):
    """
    Vista para iniciar una nueva evaluación de un negociador
    """
    if request.user.role != 'lider':
        return redirect('profile')
    
    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    
    
    if request.method == 'POST':
        form = EvaluationForm(request.POST)
        if form.is_valid():
            # Crear la evaluación
            # Crear la evaluación
            # Calcular la puntuación general automáticamente
            overall_score = negotiator.calcular_puntuacion_hacer()
            evaluation = Evaluation.objects.create(
                negotiator=negotiator,
                evaluator=request.user,
                overall_score=overall_score if overall_score is not None else 0.0,
                feedback=form.cleaned_data['feedback']
            )

            # Crear EvaluationKPI para cada KPI de tipo porcentaje usando el indicador histórico más reciente
            latest_indicator = negotiator.indicators.order_by('-date').first()
            if latest_indicator:
                kpis = KPI.objects.filter(kpi_type='percentage')
                for kpi in kpis:
                    # Mapear el nombre del KPI al campo del indicador
                    kpi_field_map = {
                        'Conversión de Ventas': 'conversion_de_ventas',
                        'Porcentajes de Cumplimiento de Recaudo': 'porcentajes_cumplimiento_recaudo',
                        'Porcentaje de Cumplimiento de Conversión': 'porcentaje_cumplimiento_conversion',
                        'Porcentaje de Caídas de Acuerdos': 'porcentaje_caidas_acuerdos',
                    }
                    field_name = kpi_field_map.get(kpi.name)
                    if field_name and hasattr(latest_indicator, field_name):
                        value = getattr(latest_indicator, field_name)
                        EvaluationKPI.objects.create(
                            evaluation=evaluation,
                            kpi=kpi,
                            score=value
                        )
            
            
            messages.success(request, f'Evaluación creada exitosamente para {negotiator.name}.')
            return redirect('negotiator_detail', cedula=cedula)
    else:
        form = EvaluationForm()
    
    context = {
        'negotiator': negotiator,
        'form': form
    }
    return render(request, 'accounts/start_evaluation.html', context)

@login_required
def negotiator_indicators_view(request, cedula):
    """
    Vista para mostrar indicadores históricos y actuales del negociador
    """
    if request.user.role != 'lider':
        return redirect('profile')
    
    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    
    # Obtener indicadores de los últimos 6 meses
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=180)
    
    indicators = negotiator.indicators.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    # Obtener indicador más reciente
    latest_indicator = negotiator.indicators.order_by('-date').first()
    
    context = {
        'negotiator': negotiator,
        'indicators': indicators,
        'latest_indicator': latest_indicator,
        'has_indicators': indicators.exists()
    }
    return render(request, 'accounts/negotiator_indicators.html', context)

@login_required
def negotiator_indicators_api(request, cedula):
    """
    API endpoint para obtener datos de indicadores en formato JSON para los gráficos
    """
    if request.user.role != 'lider':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    
    # Obtener indicadores de los últimos 6 meses
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=180)
    
    indicators = negotiator.indicators.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    # Preparar datos para los gráficos con los nuevos KPIs
    data = {
        'labels': [],
        'conversion_de_ventas': [],
        'recaudacion_mensual': [],
        'tiempo_hablando': [],
        'porcentajes_cumplimiento_recaudo': [],
        'porcentaje_cumplimiento_conversion': [],
        'porcentaje_caidas_acuerdos': []
    }

    for indicator in indicators:
        data['labels'].append(indicator.date.strftime('%Y-%m-%d'))
        data['conversion_de_ventas'].append(indicator.conversion_de_ventas)
        data['recaudacion_mensual'].append(indicator.recaudacion_mensual)
        data['tiempo_hablando'].append(indicator.tiempo_hablando)
        data['porcentajes_cumplimiento_recaudo'].append(indicator.porcentajes_cumplimiento_recaudo)
        data['porcentaje_cumplimiento_conversion'].append(indicator.porcentaje_cumplimiento_conversion)
        data['porcentaje_caidas_acuerdos'].append(indicator.porcentaje_caidas_acuerdos)
    
    return JsonResponse(data)

@login_required
def administrativo_dashboard_view(request):
    # Solo roles administrativos o superusuarios
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    # Parámetros de filtro: rango personalizado (desde/hasta) o año/semestre
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    year = None
    semestre = None
    start_date = None
    end_date = None

    # Si vienen fechas explícitas, tienen prioridad
    if desde_str or hasta_str:
        try:
            if desde_str:
                start_date = datetime.strptime(desde_str, '%Y-%m-%d').date()
            if hasta_str:
                end_date = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = None
            end_date = None

        # Defaults si falta alguno
        if start_date is None:
            start_date = timezone.now().date() - timedelta(days=180)
        if end_date is None:
            end_date = timezone.now().date()

        # Asegurar orden correcto
        if end_date < start_date:
            start_date, end_date = end_date, start_date

        # Set valores mostrados en formulario auxiliar
        year = start_date.year
        semestre = '1' if start_date.month <= 6 else '2'
    else:
        # Filtro por año/semestre (por defecto)
        try:
            year = int(request.GET.get('anio') or timezone.now().year)
        except ValueError:
            year = timezone.now().year
        semestre = request.GET.get('semestre') or '1'
        if semestre not in ['1', '2']:
            semestre = '1'

        if semestre == '1':
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 6, 30).date()
        else:
            start_date = datetime(year, 7, 1).date()
            end_date = datetime(year, 12, 31).date()

    # Consolidar KPIs por líder en el rango
    qs = (
        NegotiatorIndicator.objects
        .filter(date__gte=start_date, date__lte=end_date)
        .values(
            'negotiator__leader__cedula',
            'negotiator__leader__first_name',
            'negotiator__leader__last_name',
            'negotiator__leader__email'
        )
        .annotate(
            equipos=Count('negotiator', distinct=True),
            avg_conversion=Avg('conversion_de_ventas'),
            avg_recaudo=Avg('recaudacion_mensual'),
            avg_tiempo=Avg('tiempo_hablando'),
            avg_cump_recaudo=Avg('porcentajes_cumplimiento_recaudo'),
            avg_cump_conv=Avg('porcentaje_cumplimiento_conversion'),
            avg_caidas=Avg('porcentaje_caidas_acuerdos'),
        )
    )

    # Umbral configurable desde settings (se usa también para alertas por líder)
    semester_target = getattr(settings, 'SEMESTER_TARGET', 70)

    # Threshold to show a warning color (e.g., 80% of the target)
    warning_threshold = round(semester_target * 0.8, 2)

    leaders_data = []
    for row in qs:
        desempeno_componentes = []
        if row['avg_conversion'] is not None:
            desempeno_componentes.append(row['avg_conversion'])
        if row['avg_cump_recaudo'] is not None:
            desempeno_componentes.append(row['avg_cump_recaudo'])
        if row['avg_cump_conv'] is not None:
            desempeno_componentes.append(row['avg_cump_conv'])
        if row['avg_caidas'] is not None:
            desempeno_componentes.append(max(0.0, 100.0 - row['avg_caidas']))
        desempeno = round(sum(desempeno_componentes) / len(desempeno_componentes), 2) if desempeno_componentes else None
        leaders_data.append({
            'cedula': row['negotiator__leader__cedula'],
            'nombre': f"{row['negotiator__leader__first_name']} {row['negotiator__leader__last_name']}",
            'email': row['negotiator__leader__email'],
            'equipos': row['equipos'],
            'avg_conversion': row['avg_conversion'],
            'avg_recaudo': row['avg_recaudo'],
            'avg_tiempo': row['avg_tiempo'],
            'avg_cump_recaudo': row['avg_cump_recaudo'],
            'avg_cump_conv': row['avg_cump_conv'],
            'avg_caidas': row['avg_caidas'],
            'desempeno': desempeno,
            # Placeholder for negociadores; will be filled below
            'negociadores': []
        })

    # Ordenamiento
    ordenar_por = request.GET.get('ordenar_por') or 'desempeno'
    direccion = request.GET.get('direccion') or 'desc'
    reverse = True if direccion == 'desc' else False
    leaders_data.sort(key=lambda x: (x[ordenar_por] is None, x[ordenar_por]), reverse=reverse)

    # Enriquecer cada líder con su lista de negociadores y métricas agregadas en el rango
    for leader in leaders_data:
        try:
            negotiators_qs = Negotiator.objects.filter(leader__cedula=leader['cedula'])
        except Exception:
            negotiators_qs = Negotiator.objects.filter(leader__cedula=leader.get('cedula'))

        negociadores_list = []
        for n in negotiators_qs:
            # Agregar promedios desde NegotiatorIndicator en el mismo rango de fechas
            ind_qs = n.indicators.filter(date__gte=start_date, date__lte=end_date)
            agg = ind_qs.aggregate(
                avg_conversion=Coalesce(Avg('conversion_de_ventas'), 0.0),
                avg_recaudo=Coalesce(Avg('recaudacion_mensual'), 0.0),
                avg_tiempo=Coalesce(Avg('tiempo_hablando'), 0.0),
                avg_cump_recaudo=Coalesce(Avg('porcentajes_cumplimiento_recaudo'), 0.0),
                avg_cump_conv=Coalesce(Avg('porcentaje_cumplimiento_conversion'), 0.0),
                avg_caidas=Coalesce(Avg('porcentaje_caidas_acuerdos'), 0.0),
            )
            # Calcular desempeño similar al líder
            componentes = []
            if agg['avg_conversion'] is not None:
                componentes.append(agg['avg_conversion'])
            if agg['avg_cump_recaudo'] is not None:
                componentes.append(agg['avg_cump_recaudo'])
            if agg['avg_cump_conv'] is not None:
                componentes.append(agg['avg_cump_conv'])
            if agg['avg_caidas'] is not None:
                componentes.append(max(0.0, 100.0 - agg['avg_caidas']))
            n_desempeno = round(sum(componentes) / len(componentes), 2) if componentes else None

            negociadores_list.append({
                'cedula': n.cedula,
                'nombre': n.name,
                'avg_conversion': agg.get('avg_conversion'),
                'avg_recaudo': agg.get('avg_recaudo'),
                'avg_tiempo': agg.get('avg_tiempo'),
                'avg_cump_recaudo': agg.get('avg_cump_recaudo'),
                'avg_cump_conv': agg.get('avg_cump_conv'),
                'avg_caidas': agg.get('avg_caidas'),
                'desempeno': n_desempeno,
                # Indica si este negociador tuvo al menos una evaluación en el rango seleccionado
                'evaluated_in_range': n.evaluations.filter(date__date__gte=start_date, date__date__lte=end_date).exists(),
            })

        leader['negociadores'] = negociadores_list

        # --- Métricas de diligenciamiento por líder (HU20) ---
        # Total negociadores del líder (fuente primaria)
        total_negotiators = Negotiator.objects.filter(leader__cedula=leader['cedula']).count()

        # Número de negociadores que tuvieron al menos una evaluación en el rango
        evaluated_negotiators = Negotiator.objects.filter(
            leader__cedula=leader['cedula'],
            evaluations__date__date__gte=start_date,
            evaluations__date__date__lte=end_date
        ).distinct().count()

        # Número total de evaluaciones registradas para el líder en el rango
        evaluations_done = Evaluation.objects.filter(
            negotiator__leader__cedula=leader['cedula'],
            date__date__gte=start_date,
            date__date__lte=end_date
        ).count()

        pending = total_negotiators - evaluated_negotiators
        pct = round((evaluated_negotiators / total_negotiators) * 100, 2) if total_negotiators > 0 else 0

        leader['total_negotiators'] = total_negotiators
        leader['evaluated_negotiators'] = evaluated_negotiators
        leader['evaluations_done'] = evaluations_done
        leader['pending_negotiators'] = pending
        leader['pct_cumplimiento'] = pct
        # Flag para alerta visual (por debajo del objetivo semestral)
        leader['alert'] = pct < semester_target if total_negotiators > 0 else False

    # Preparar datos para los gráficos (JSON serializable)
    labels = [l['nombre'] for l in leaders_data]
    counts = [int(l.get('equipos') or 0) for l in leaders_data]
    total_counts = sum(counts)
    shares = [round((c / total_counts) * 100, 2) if total_counts > 0 else 0 for c in counts]
    avg_desempeno = [round(l['desempeno'], 2) if l.get('desempeno') is not None else 0 for l in leaders_data]

    # KPI arrays (aligned with labels)
    chart_avg_recaudo = [round(l.get('avg_recaudo') or 0, 2) for l in leaders_data]
    chart_avg_tiempo = [round(l.get('avg_tiempo') or 0, 2) for l in leaders_data]
    chart_avg_conversion = [round(l.get('avg_conversion') or 0, 2) for l in leaders_data]
    chart_avg_cump_recaudo = [round(l.get('avg_cump_recaudo') or 0, 2) for l in leaders_data]
    chart_avg_cump_conv = [round(l.get('avg_cump_conv') or 0, 2) for l in leaders_data]
    chart_avg_caidas = [round(l.get('avg_caidas') or 0, 2) for l in leaders_data]

    # Calcular desempeño global (promedio de líderes con valor definido)
    non_null_desempenos = [l['desempeno'] for l in leaders_data if l.get('desempeno') is not None]
    overall_desempeno = round(sum(non_null_desempenos) / len(non_null_desempenos), 2) if non_null_desempenos else None

    # Umbral configurable desde settings
    semester_target = getattr(settings, 'SEMESTER_TARGET', 70)

    context = {
        'leaders_data': leaders_data,
        'anio': year,
        'semestre': semestre,
        'ordenar_por': ordenar_por,
        'direccion': direccion,
        'start_date': start_date,
        'end_date': end_date,
        'desde': start_date.strftime('%Y-%m-%d') if start_date else None,
        'hasta': end_date.strftime('%Y-%m-%d') if end_date else None,

        # Gráficos (pasados como JSON seguro)
        'chart_labels': json.dumps(labels),
        'chart_counts': json.dumps(counts),
        'chart_shares': json.dumps(shares),
        'chart_avg_desempeno': json.dumps(avg_desempeno),

        'chart_avg_recaudo': json.dumps(chart_avg_recaudo),
        'chart_avg_tiempo': json.dumps(chart_avg_tiempo),
        'chart_avg_conversion': json.dumps(chart_avg_conversion),
        'chart_avg_cump_recaudo': json.dumps(chart_avg_cump_recaudo),
        'chart_avg_cump_conv': json.dumps(chart_avg_cump_conv),
        'chart_avg_caidas': json.dumps(chart_avg_caidas),
        'overall_desempeno': overall_desempeno,
        'semester_target': semester_target,
        'warning_threshold': warning_threshold,
    }
    return render(request, 'accounts/administrativo_dashboard.html', context)


@login_required
def admin_evaluation_detail(request, pk):
    """Vista de solo lectura para una evaluación (Hacer) accesible solo por administradores."""
    # Permitir solo administradores o superusuarios
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    evaluation = get_object_or_404(Evaluation, pk=pk)

    if request.method == 'POST':
        # No permitimos modificaciones: informar y mostrar vista de solo lectura
        messages.warning(request, 'Esta evaluación es de solo lectura para administradores y no puede modificarse aquí.')

    # KPIs asociados a la evaluación (EvaluationKPI)
    evaluation_kpis_qs = evaluation.kpis.select_related('kpi').all()
    # Pre-compute display values to avoid calling model methods with args in templates
    evaluation_kpis = []
    for ek in evaluation_kpis_qs:
        try:
            display = ek.kpi.get_display_value(ek.score)
        except Exception:
            display = ek.score
        evaluation_kpis.append({
            'kpi_name': ek.kpi.name,
            'score': ek.score,
            'display': display,
        })

    # Buscar el indicador más cercano anterior o igual a la fecha de la evaluación
    latest_indicator = None
    try:
        eval_date = evaluation.date.date()
        latest_indicator = evaluation.negotiator.indicators.filter(date__lte=eval_date).order_by('-date').first()
    except Exception:
        latest_indicator = None

    context = {
        'evaluation': evaluation,
        'evaluation_kpis': evaluation_kpis,
        'latest_indicator': latest_indicator,
    }
    return render(request, 'accounts/admin_evaluation_detail.html', context)


@login_required
def admin_ser_evaluation_detail(request, pk):
    """Vista de solo lectura para una evaluación del Ser accesible solo por administradores."""
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    ser_eval = get_object_or_404(SerEvaluation, pk=pk)

    if request.method == 'POST':
        messages.warning(request, 'Esta evaluación (Ser) es de solo lectura para administradores y no puede modificarse aquí.')

    context = {
        'ser_eval': ser_eval,
    }
    return render(request, 'accounts/admin_ser_evaluation_detail.html', context)

@login_required
def exportar_resultados_excel(request):
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    # Soporta exportación por fechas (desde/hasta) o por año/semestre (backward compatible)
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    start_date = None
    end_date = None
    if desde_str or hasta_str:
        try:
            if desde_str:
                start_date = datetime.strptime(desde_str, '%Y-%m-%d').date()
            if hasta_str:
                end_date = datetime.strptime(hasta_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = None
            end_date = None
        if start_date is None:
            start_date = timezone.now().date() - timedelta(days=180)
        if end_date is None:
            end_date = timezone.now().date()
        if end_date < start_date:
            start_date, end_date = end_date, start_date
    else:
        try:
            year = int(request.GET.get('anio') or timezone.now().year)
        except ValueError:
            year = timezone.now().year
        semestre = request.GET.get('semestre') or '1'
        if semestre not in ['1', '2']:
            semestre = '1'
        if semestre == '1':
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 6, 30).date()
        else:
            start_date = datetime(year, 7, 1).date()
            end_date = datetime(year, 12, 31).date()

    # Datos consolidados
    qs = (
        NegotiatorIndicator.objects
        .filter(date__gte=start_date, date__lte=end_date)
        .values(
            'negotiator__leader__cedula',
            'negotiator__leader__first_name',
            'negotiator__leader__last_name',
            'negotiator__leader__email'
        )
        .annotate(
            equipos=Count('negotiator', distinct=True),
            avg_conversion=Avg('conversion_de_ventas'),
            avg_recaudo=Avg('recaudacion_mensual'),
            avg_tiempo=Avg('tiempo_hablando'),
            avg_cump_recaudo=Avg('porcentajes_cumplimiento_recaudo'),
            avg_cump_conv=Avg('porcentaje_cumplimiento_conversion'),
            avg_caidas=Avg('porcentaje_caidas_acuerdos'),
        )
    )

    # Generar Excel con openpyxl
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse('Falta dependencia openpyxl. Instálala e inténtalo de nuevo.', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Semestre {semestre} {year}'
    headers = [
        'Cédula Líder', 'Nombre Líder', 'Email', 'Negociadores',
        'Conv. Ventas (%)', 'Recaudo ($)', 'Tiempo Hablando (h)',
        'Cumpl. Recaudo (%)', 'Cumpl. Conversión (%)', 'Caídas Acuerdos (%)', 'Desempeño'
    ]
    ws.append(headers)

    for row in qs:
        desempeno_componentes = []
        if row['avg_conversion'] is not None:
            desempeno_componentes.append(row['avg_conversion'])
        if row['avg_cump_recaudo'] is not None:
            desempeno_componentes.append(row['avg_cump_recaudo'])
        if row['avg_cump_conv'] is not None:
            desempeno_componentes.append(row['avg_cump_conv'])
        if row['avg_caidas'] is not None:
            desempeno_componentes.append(max(0.0, 100.0 - row['avg_caidas']))
        desempeno = round(sum(desempeno_componentes) / len(desempeno_componentes), 2) if desempeno_componentes else None

        nombre = f"{row['negotiator__leader__first_name']} {row['negotiator__leader__last_name']}"
        ws.append([
            row['negotiator__leader__cedula'],
            nombre,
            row['negotiator__leader__email'],
            row['equipos'],
            _round(row['avg_conversion']),
            _round(row['avg_recaudo']),
            _round(row['avg_tiempo']),
            _round(row['avg_cump_recaudo']),
            _round(row['avg_cump_conv']),
            _round(row['avg_caidas']),
            _round(desempeno),
        ])

    # Ajuste simple de ancho
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 20

    # Respuesta HTTP
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    # Nombre de archivo sensible al filtro utilizado
    if desde_str or hasta_str:
        filename = f'resultados_{start_date}_{end_date}.xlsx'
    else:
        filename = f'resultados_semestre_{semestre}_{year}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _round(val):
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except Exception:
        return val

@login_required
def exportar_evaluacion_pdf(request, cedula):
    # Solo líderes pueden exportar sus evaluaciones
    if request.user.role != 'lider':
        return redirect('profile')

    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    last_eval = negotiator.evaluations.order_by('-date').first()
    last_ser = negotiator.ser_evaluations.order_by('-date').first()

    if not last_eval:
        messages.warning(request, 'No hay evaluación para exportar.')
        return redirect('negotiator_detail', cedula=cedula)

    # Datos KPI de la evaluación
    kpis = list(last_eval.kpis.select_related('kpi').all())

    # Construir PDF con reportlab
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return HttpResponse('Falta dependencia reportlab. Instálala e inténtalo de nuevo.', status=500)

    from io import BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    story = []
    story.append(Paragraph('Evaluación Final con Retroalimentación', styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f'Negociador: <b>{negotiator.name}</b> (Cédula: {negotiator.cedula})', styles['Normal']))
    story.append(Paragraph(f'Líder: <b>{request.user.first_name} {request.user.last_name}</b> ({request.user.email})', styles['Normal']))
    story.append(Paragraph(f'Fecha: {last_eval.date.strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    story.append(Spacer(1, 12))

    # Tabla KPIs
    data = [['Indicador', 'Valor (%)']]
    for ek in kpis:
        data.append([ek.kpi.name, f"{_round(ek.score)}"])
    t = Table(data, hAlign='LEFT')
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    # Porcentajes generales
    hacer_pct = _round(last_eval.overall_score)
    ser_pct = _round((last_ser.promedio * 20) if last_ser else None)
    total_pct = None
    if hacer_pct is not None and ser_pct is not None:
        total_pct = _round(hacer_pct * 0.7 + ser_pct * 0.3)

    resumen = [['Hacer (0-100)', hacer_pct], ['Ser (x20 → 0-100)', ser_pct], ['Total (70/30)', total_pct]]
    t2 = Table(resumen, hAlign='LEFT')
    t2.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
    ]))
    story.append(t2)
    story.append(Spacer(1, 12))

    # Retroalimentación
    if last_eval.feedback:
        story.append(Paragraph('Retroalimentación', styles['Heading3']))
        story.append(Paragraph(last_eval.feedback, styles['BodyText']))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="evaluacion_{negotiator.cedula}.pdf"'
    return response

@login_required
def negotiator_detail_view(request, cedula):
    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)
    last_evaluation = Evaluation.objects.filter(negotiator=negotiator).order_by('-date').first()
    context = {
        'negotiator': negotiator,
        'last_evaluation': last_evaluation
    }
    return render(request, 'accounts/negotiator_detail.html', context)


@login_required
def generar_sugerencia_view(request, cedula):
    if request.user.role != 'lider':
        return redirect('profile')

    negotiator = get_object_or_404(Negotiator, cedula=cedula, leader=request.user)

    latest_indicator = negotiator.indicators.order_by('-date').first()
    last_eval = negotiator.evaluations.order_by('-date').first()
    last_ser = negotiator.ser_evaluations.order_by('-date').first()

    if not latest_indicator and not last_eval and not last_ser:
        messages.warning(request, 'No hay datos suficientes para generar una sugerencia.')
        return redirect('negotiator_detail', cedula=cedula)

    suggestions = []

    if latest_indicator:
        try:
            if latest_indicator.porcentaje_cumplimiento_conversion is not None and latest_indicator.porcentaje_cumplimiento_conversion < 60:
                suggestions.append('Reforzar el seguimiento de leads y la calidad del discurso para mejorar el cumplimiento de conversión (<60%).')
        except Exception:
            pass

        try:
            if latest_indicator.porcentajes_cumplimiento_recaudo is not None and latest_indicator.porcentajes_cumplimiento_recaudo < 70:
                suggestions.append('Aumentar la cadencia de recordatorios y acuerdos de pago para mejorar el recaudo (<70%).')
        except Exception:
            pass

        try:
            if latest_indicator.porcentaje_caidas_acuerdos is not None and latest_indicator.porcentaje_caidas_acuerdos > 20:
                suggestions.append('Revisar objeciones frecuentes y fortalecer cierres para reducir caídas de acuerdos (>20%).')
        except Exception:
            pass

        try:
            if latest_indicator.conversion_de_ventas is not None and latest_indicator.conversion_de_ventas < 30:
                suggestions.append('Trabajar el pitch inicial y calificación de oportunidades: la conversión de ventas está por debajo de 30%.')
        except Exception:
            pass

    if last_eval and last_eval.overall_score is not None and last_eval.overall_score < 60:
        suggestions.append('Foco en KPIs del Hacer con menor desempeño del último mes para elevar el puntaje general (<60).')

    if last_ser and last_ser.promedio is not None and last_ser.promedio < 3.5:
        suggestions.append('Refuerzo en habilidades blandas (trabajo en equipo, comunicación y compromiso) para mejorar el SER (<3.5/5).')

    if not suggestions:
        suggestions.append('Buen desempeño general. Mantener prácticas actuales y compartir mejores prácticas con el equipo.')

    messages.info(request, 'Sugerencia: ' + ' | '.join(suggestions))
    return redirect('negotiator_detail', cedula=cedula)


# ============================
# Histórico de Evaluaciones
# ============================
from django.contrib.auth import get_user_model


@login_required
def historico_evaluaciones_view(request):
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    # Filtros
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    lider_cedula = request.GET.get('lider')
    negociador_cedula = request.GET.get('negociador')

    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else None
    except ValueError:
        desde = None
    try:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else None
    except ValueError:
        hasta = None
    if desde and hasta and hasta < desde:
        desde, hasta = hasta, desde

    evals = Evaluation.objects.select_related('negotiator', 'evaluator').all()
    ser_evals = SerEvaluation.objects.select_related('negotiator', 'evaluator').all()

    if desde:
        evals = evals.filter(date__date__gte=desde)
        ser_evals = ser_evals.filter(date__date__gte=desde)
    if hasta:
        evals = evals.filter(date__date__lte=hasta)
        ser_evals = ser_evals.filter(date__date__lte=hasta)
    if lider_cedula:
        evals = evals.filter(evaluator__cedula=lider_cedula)
        ser_evals = ser_evals.filter(evaluator__cedula=lider_cedula)
    if negociador_cedula:
        evals = evals.filter(negotiator__cedula=negociador_cedula)
        ser_evals = ser_evals.filter(negotiator__cedula=negociador_cedula)

    # Listas auxiliares para selects
    User = get_user_model()
    leaders = User.objects.filter(role='lider').order_by('first_name', 'last_name').values('cedula', 'first_name', 'last_name', 'email')
    negotiators = Negotiator.objects.order_by('name').values('cedula', 'name')

    # Estadísticas generales
    from django.db.models import Avg
    avg_hacer = evals.aggregate(avg=Avg('overall_score'))['avg']
    # promedio SER calculado en Python para no complicar el ORM con anotaciones
    ser_proms = [se.promedio for se in ser_evals]
    avg_ser = round(sum(ser_proms) / len(ser_proms), 2) if ser_proms else None

    context = {
        'evals': evals.order_by('-date'),
        'ser_evals': ser_evals.order_by('-date'),
        'leaders': list(leaders),
        'negotiators': list(negotiators),
        'desde': desde.strftime('%Y-%m-%d') if desde else '',
        'hasta': hasta.strftime('%Y-%m-%d') if hasta else '',
        'lider_selected': lider_cedula or '',
        'negociador_selected': negociador_cedula or '',
        'avg_hacer': _round(avg_hacer) if avg_hacer is not None else None,
        'avg_ser': _round(avg_ser) if avg_ser is not None else None,
        'count_hacer': evals.count(),
        'count_ser': ser_evals.count(),
    }
    return render(request, 'accounts/historico_evaluaciones.html', context)


@login_required
def exportar_historico_excel(request):
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    lider_cedula = request.GET.get('lider')
    negociador_cedula = request.GET.get('negociador')

    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else None
    except ValueError:
        desde = None
    try:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else None
    except ValueError:
        hasta = None
    if desde and hasta and hasta < desde:
        desde, hasta = hasta, desde

    evals = Evaluation.objects.select_related('negotiator', 'evaluator').all()
    ser_evals = SerEvaluation.objects.select_related('negotiator', 'evaluator').all()

    if desde:
        evals = evals.filter(date__date__gte=desde)
        ser_evals = ser_evals.filter(date__date__gte=desde)
    if hasta:
        evals = evals.filter(date__date__lte=hasta)
        ser_evals = ser_evals.filter(date__date__lte=hasta)
    if lider_cedula:
        evals = evals.filter(evaluator__cedula=lider_cedula)
        ser_evals = ser_evals.filter(evaluator__cedula=lider_cedula)
    if negociador_cedula:
        evals = evals.filter(negotiator__cedula=negociador_cedula)
        ser_evals = ser_evals.filter(negotiator__cedula=negociador_cedula)

    try:
        import openpyxl
    except Exception:
        return HttpResponse('Falta dependencia openpyxl. Instálala e inténtalo de nuevo.', status=500)

    wb = openpyxl.Workbook()
    # Hoja Hacer
    ws1 = wb.active
    ws1.title = 'Evaluaciones Hacer'
    ws1.append(['Fecha', 'Líder', 'Correo Líder', 'Negociador', 'Cédula Negociador', 'Puntaje Hacer (0-100)', 'Feedback'])
    for e in evals.order_by('-date'):
        ws1.append([
            e.date.strftime('%Y-%m-%d %H:%M'),
            f'{e.evaluator.first_name} {e.evaluator.last_name}',
            e.evaluator.email,
            e.negotiator.name,
            e.negotiator.cedula,
            _round(e.overall_score),
            e.feedback or ''
        ])

    # Hoja Ser
    ws2 = wb.create_sheet('Evaluaciones Ser')
    ws2.append(['Fecha', 'Líder', 'Correo Líder', 'Negociador', 'Cédula Negociador', 'Actitud', 'Trabajo en Equipo', 'Sentido de Pertenencia', 'Relacionamiento', 'Compromiso', 'Promedio (1-5)'])
    for s in ser_evals.order_by('-date'):
        ws2.append([
            s.date.strftime('%Y-%m-%d %H:%M'),
            f'{s.evaluator.first_name} {s.evaluator.last_name}',
            s.evaluator.email,
            s.negotiator.name,
            s.negotiator.cedula,
            s.actitud,
            s.trabajo_en_equipo,
            s.sentido_pertenencia,
            s.relacionamiento,
            s.compromiso,
            s.promedio,
        ])

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = 'historico_evaluaciones.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def exportar_historico_pdf(request):
    if request.user.role != 'administrativo' and not request.user.is_superuser:
        return redirect('profile')

    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    lider_cedula = request.GET.get('lider')
    negociador_cedula = request.GET.get('negociador')

    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else None
    except ValueError:
        desde = None
    try:
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else None
    except ValueError:
        hasta = None
    if desde and hasta and hasta < desde:
        desde, hasta = hasta, desde

    evals = Evaluation.objects.select_related('negotiator', 'evaluator').all()
    ser_evals = SerEvaluation.objects.select_related('negotiator', 'evaluator').all()

    if desde:
        evals = evals.filter(date__date__gte=desde)
        ser_evals = ser_evals.filter(date__date__gte=desde)
    if hasta:
        evals = evals.filter(date__date__lte=hasta)
        ser_evals = ser_evals.filter(date__date__lte=hasta)
    if lider_cedula:
        evals = evals.filter(evaluator__cedula=lider_cedula)
        ser_evals = ser_evals.filter(evaluator__cedula=lider_cedula)
    if negociador_cedula:
        evals = evals.filter(negotiator__cedula=negociador_cedula)
        ser_evals = ser_evals.filter(negotiator__cedula=negociador_cedula)

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse('Falta dependencia reportlab. Instálala e inténtalo de nuevo.', status=500)

    from io import BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    # Título
    title = 'Histórico de Evaluaciones'
    subt = []
    if desde:
        subt.append(f'Desde: {desde.strftime("%Y-%m-%d")}')
    if hasta:
        subt.append(f'Hasta: {hasta.strftime("%Y-%m-%d")}')
    story.append(Paragraph(title, styles['Title']))
    if subt:
        story.append(Paragraph(' | '.join(subt), styles['Normal']))
    story.append(Spacer(1, 12))

    # Tabla Hacer
    data_h = [['Fecha', 'Líder', 'Negociador', 'Hacer (0-100)']]
    for e in evals.order_by('-date')[:500]:  # limitar filas para no crecer demasiado
        data_h.append([
            e.date.strftime('%Y-%m-%d %H:%M'),
            f'{e.evaluator.first_name} {e.evaluator.last_name}',
            e.negotiator.name,
            _round(e.overall_score)
        ])
    t_h = Table(data_h, hAlign='LEFT')
    t_h.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(Paragraph('Evaluaciones Hacer', styles['Heading2']))
    story.append(t_h)
    story.append(Spacer(1, 12))

    # Tabla Ser
    data_s = [['Fecha', 'Líder', 'Negociador', 'Promedio SER (1-5)']]
    for s in ser_evals.order_by('-date')[:500]:
        data_s.append([
            s.date.strftime('%Y-%m-%d %H:%M'),
            f'{s.evaluator.first_name} {s.evaluator.last_name}',
            s.negotiator.name,
            s.promedio
        ])
    t_s = Table(data_s, hAlign='LEFT')
    t_s.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    story.append(Paragraph('Evaluaciones Ser', styles['Heading2']))
    story.append(t_s)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="historico_evaluaciones.pdf"'
    return response
